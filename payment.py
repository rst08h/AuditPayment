import screen_config
import tkinter as tk
from tkinter import filedialog as fd
import pandas as pd
import threading
import time
from datetime import datetime
import misc
from decimal import Decimal
from decimal import getcontext
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from sys import platform
import os


#from tkinter.scrolledtext import ScrolledText

getcontext().prec = 2

###################################################################################
#
#   PROCESS FILE 1 ไฟล์ ZFAPRP08 โดย User export มาเป็น .xls แต่จริงๆ ข้าในเป็น TEXT FILE utf_16_le
#
###################################################################################
def process_file():

    ## รับค่า path ของไฟล์ต้นทาง
    if filename1.cget('text') != "[ไม่ได้เลือก]":
        result_path=filename1.cget('text')
    elif filename2.cget('text') != "[ไม่ได้เลือก]":
        result_path=filename2.cget('text')
    else:
        result_path=''

    result_path=os.path.dirname(result_path)

    tabtype=0
    ZFAPRP08_amount = None
    gl_amount = None
    pwabranch = None
    if len(txtbox1.get("1.0", tk.END)) >= 2:
        txtbox1.delete("1.0", tk.END)

    if filename1.cget('text') != "[ไม่ได้เลือก]":
        txtbox1.insert(tk.END, '[ประมวลผลไฟล์ ZFAPRP08]\n')
        txtbox1.yview_pickplace("end")
        saptextfile = filename1.cget('text')
        saptextfile2 = saptextfile+'.txt'
        file1 = open(saptextfile, 'r', encoding='utf_16_le')
        file2 = open(saptextfile2, 'w', encoding='utf-8')
        file1.read(1)
        line = file1.readline()
        if line.find('ZFAPRP08') > 0:
            txtbox1.insert(tk.END, 'อ่านไฟล์ ' + saptextfile + '\n')
            txtbox1.yview_pickplace("end")
            while line != '':
                # print(line)
                if (line[0] != '0') and (line[0] != '1') and (line[0] != '2') and (line[0] != '3') and (line[0] != '`'):
                    file2.write('`'+line)
                    if pwabranch == None:
                        if line.find('กปภ.') > 0:
                            pwabranch = line[line.find('กปภ.'):-1].rstrip()
                            txtbox1.insert(
                                tk.END, 'หน่วยรับตรวจ: ' + pwabranch + '\n')
                            txtbox1.yview_pickplace("end")

                else:
                    #line=misc.line_clinsing(line,tabtype)
                    line=misc.line_clinsing(line)
                    file2.write(line)
                    print(line.count('\t'))
                line = file1.readline()
            file1.close()
            file2.close()
##################################################
#
#   อ่านไฟล์ ZF
#
##################################################
            def dateparse(dates): return [
                datetime.strptime(d, '%d.%m.%Y') for d in dates]
            cols = [0, 1, 2, 3, 4, 5, 6, 7]
            colsname = ['posting_date', 'doc_no', 'tax_no', 'doctype',
                        'payee_name', 'description', 'amount', 'collected']

            df1 = pd.read_table(saptextfile2, sep='\t', engine='python', date_parser=dateparse, parse_dates=[
                'posting_date'], header=None, usecols=cols, names=colsname, comment='`', dtype={'doctype': 'category', 'collected': 'category'}, encoding='utf-8')
            df1.doc_no = df1.doc_no.astype('str')
            #df1['amount'] = df1['amount'].str.replace(',', '')
            df1.sort_values('posting_date')

###########################################################################################
#
#   สร้างไฟล์กระดาษทำการ หน้าสรุป
#
#
###########################################################################################
            dmin = df1['posting_date'].min()
            dmax = df1['posting_date'].max()

            thsarabun= Font('TH Sarabun New',16)
            border = Border(left=Side(color='00000000',border_style='thin'),right=Side(color='00000000',border_style='thin'),top=Side(color='00000000',border_style='thin'),bottom=Side(color='00000000',border_style='thin'))
            

            

            result_filename = result_path +  '/กระดาษทำการจ่ายเงิน ' + pwabranch +'_'+misc.thaidate(dmin)+'-'+misc.thaidate(dmax) + '.xlsx'
            xlwb = Workbook()
            # xlsheet=xlwb.create_sheet('สรุป')
            xlsheet = xlwb.active
            xlsheet.title = 'สรุป'

            for x in range(1,10):
                for y in range(1,100):
                    xlsheet.cell(y,x).font=thsarabun
            for x in range(4,6):
                for y in range(1,9):
                    xlsheet.cell(row=x,column=y).border=border



            xlsheet.column_dimensions['A'].width =11.33
            xlsheet.column_dimensions['B'].width=23
            xlsheet.column_dimensions['C'].width=23
            xlsheet.column_dimensions['D'].width=15
            xlsheet.column_dimensions['E'].width=30
            xlsheet.column_dimensions['F'].width=30
            xlsheet.column_dimensions['G'].width=30
            xlsheet.column_dimensions['H'].width=30

            
            xlsheet.merge_cells('B4:D4')
            xlsheet.merge_cells('E4:F4')
            xlsheet.merge_cells('A4:A5')
            xlsheet.merge_cells('G4:G5')
            xlsheet.merge_cells('H4:H5')
            xlsheet.cell(4,2,"ตรวจสอบการบันทึกบัญชีค่าใช้จ่ายเข้าระบบ SAP").alignment=Alignment(horizontal='center')
            xlsheet.cell(4,5,'ตรวจสอบการเบิกชดเชยเงินสดย่อยกับค่าใช้จ่ายทั้งหมดในเดือน').alignment=Alignment(horizontal='center')
            xlsheet.cell(4, 1, 'เดือน/ปี').alignment = Alignment(horizontal='center',vertical='center')
            xlsheet.cell(5, 2, 'บันทึกค่าใช้จ่ายฝั่งเดบิต').alignment=Alignment(horizontal='center')
            xlsheet.cell(5, 3, 'บันทึกค่าใช้จ่ายฝั่งเครดิต').alignment=Alignment(horizontal='center')
            xlsheet.cell(5, 4, 'ผลต่าง').alignment=Alignment(horizontal='center')
            xlsheet.cell(5, 5, 'จำนวนเงินที่เบิกชดเชยทั้งเดือน').alignment=Alignment(horizontal='center')
            xlsheet.cell(5, 6, 'ผลต่างเงินที่เบิกชดเชยกับค่าใช้จ่าย').alignment=Alignment(horizontal='center')
            xlsheet.cell(4,7,'สาเหตุผลต่าง').alignment=Alignment(horizontal='center',vertical='center')
            xlsheet.cell(4,8,'การเบิกชดเชยเงินสดย่อยเป็นไปตามหลักเกณฑ์').alignment=Alignment(horizontal='center',wrap_text=True)
            xlsheet.cell(1,7,'กระดาษทำการเลขที่ :')
            xlsheet.cell(2,7,'จัดทำโดย :')
            xlsheet.cell(3,7,'สอบทานโดย :')    

            xlsheet.cell(1,2,'หน่วยรับตรวจ : '+ pwabranch)
            xlsheet.cell(2,2,'เรื่องที่ตรวจสอบ : การเบิกชดเชยเงินสดย่อยเป็นไปตามหลักเกณฑ์')
            xlsheet.cell(3,2,'งวดตรวจสอบ : '  +misc.thaidate(dmin)+' ถึง '+misc.thaidate(dmax))
            xlwb.save(result_filename)
            xlwb.close()
            xlrow = 6

            #############
            #
            # แยกยอดเงินให้เป็นฝั่งบวกและฝั่งลบ
            df1['amount_pos'] = df1.amount[df1.amount.astype('float') >= 0]
            df1['amount_neg'] = df1.amount[df1.amount.astype('float') < 0]

            #df1['amount'] = df1['amount'].str.replace('.', '')
            # ตอนที่แสดงผลต้องหารด้วย 100
            #df1.amount = df1.amount.astype('int')
            #
            #df1.to_excel(saptextfile+'.xlsx', index=False)
            # txtbox1.insert(tk.END, 'ส่งออก ZFAPRP08 ไปยัง ' +
            #               saptextfile+'.xlsx\n')

            # เพิ่มฟิล ปีเดือน เพื่อใช้ในการกรองงวด
            df1['ym'] = df1['posting_date'].map(lambda x: x.strftime('%Y%m'))
            ym = df1.groupby('ym')['ym'].min()
            # print(misc.thaiym(df1['posting_date'].min()))
            # วิเคราะห์ข้อมูล
            txtbox1.insert(tk.END, '[วิเคราะห์ข้อมูล]\n')
            txtbox1.yview_pickplace("end")
            firstsheet = True
            for i in ym:
                df_temp = df1[df1['ym'] == i]
                print(i)
                # yyyymm=df_temp['posting_date'].min()
                tym = misc.thaiym(df_temp['posting_date'].min())
                txtbox1.insert(tk.END, 'งวดเดือน [' + tym + ']\n')
                txtbox1.yview_pickplace("end")
                # จำนวนรายการ
                record_count = df_temp['posting_date'].count()
                txtbox1.insert(tk.END, 'จำนวนรายการ : ' +
                               str(record_count) + ' รายการ\n')
                txtbox1.yview_pickplace("end")
                # ช่วงของวันที่ผ่านรายการ
                dmin = df_temp['posting_date'].min()
                dmax = df_temp['posting_date'].max()
                txtbox1.insert(tk.END, 'วันที่ผ่านรายการ ระหว่างวันที่ ' + misc.thaidate(dmin) +
                               ' ถึงวันที่ '+misc.thaidate(dmax) + '\n')
                txtbox1.yview_pickplace("end")
                # รวมฝั่งบวก และ ลบ
                positive_sum = Decimal('0')
                negative_sum = Decimal('0')
                positive_max = None
                positive_min = None
                for i in df_temp['amount']:
                    print(i)
                    if Decimal(i) > 0:
                        positive_sum += Decimal(i)
                        if positive_max != None:
                            if Decimal(i) > positive_max:
                                positive_max = Decimal(i)
                        else:
                            positive_max = Decimal(i)
                        if positive_min != None:
                            if Decimal(i) < positive_min:
                                positive_min = Decimal(i)
                        else:
                            positive_min = Decimal(i)
                    else:
                        negative_sum += Decimal(i)
                txtbox1.insert(
                    tk.END, 'จำนวนเงินฝั่งบวก = {:,.2f} บาท\n'.format(positive_sum))
                txtbox1.yview_pickplace("end")
                txtbox1.insert(
                    tk.END, 'จำนวนเงินฝั่งลบ = {:,.2f} บาท\n'.format(negative_sum))
                txtbox1.yview_pickplace("end")
                txtbox1.insert(tk.END, 'ผลต่าง {:,.2f} บาท <== ต้องเป็น 0.00\n'.format(
                    positive_sum+negative_sum))
                txtbox1.yview_pickplace("end")
                txtbox1.insert(
                    tk.END, 'จำนวนเงินต่ำสุดคือ {:,.2f} บาท\n'.format(positive_min))
                txtbox1.yview_pickplace("end")
                txtbox1.insert(
                    tk.END, 'จำนวนเงินสูงสุดคือ {:,.2f} บาท\n'.format(positive_max))
                txtbox1.yview_pickplace("end")
                #ZFAPRP08_amount[i] = positive_sum
                # convert amount from Decimal to float
                #df_temp['amount'] = df_temp['amount'].astype('float')
                df_temp['amount_pos'] = df_temp['amount_pos'].astype('float')
                df_temp['amount_neg'] = df_temp['amount_neg'].astype('float')

                #
                # Write to Excel
                #
                # เพิ่มบันทัดสรุป
                xlwb = openpyxl.load_workbook(result_filename)
                xlsheet = xlwb.get_sheet_by_name('สรุป')
                xlsheet.cell(xlrow, 1, tym)
                xlsheet.cell(xlrow, 2, "=SUM('เงินสดย่อย "+tym + "'!I:I)").number_format='#,##0.00'
                xlsheet.cell(xlrow, 3, "=SUM('เงินสดย่อย "+tym + "'!J:J)").number_format='#,##0.00'
                xlsheet.cell(xlrow, 4, "=sum(B" +str(xlrow) + ":C" +str(xlrow)+ ")").number_format='#,##0.00'
                xlsheet.cell(xlrow, 5, "=SUM('GL "+tym + "'!G:G)").number_format='#,##0.00'
                xlsheet.cell(xlrow, 6, "=B" + str(xlrow) + "+E" +str(xlrow) ).number_format='#,##0.00'
                for x in range(1,9):
                    xlsheet.cell(xlrow,x).border=border
                xlwb.save(result_filename)
                xlwb.close()
                xlrow += 1

                # เพิ่มชีตข้อมูล

                with pd.ExcelWriter(result_filename, mode='a', engine='openpyxl') as writer:
                    df_temp.to_excel(
                        writer, sheet_name='เงินสดย่อย '+tym, index=False)
                    txtbox1.insert(tk.END, 'ส่งออก ZFAPRP08 ไปยัง "' +
                                   result_filename + '"\nSheet Name: ' + 'เงินสดย่อย '+tym + '\n\n')
                    txtbox1.yview_pickplace("end")
        else:
            print("ไม่ใช่ไฟล์ ZFAPRP08")
            txtbox1.insert(tk.END, 'ไม่ใช่ไฟล์ ZFAPRP08\n')
            txtbox1.yview_pickplace("end")
        txtbox1.insert(tk.END, '\n')
        txtbox1.yview_pickplace("end")


###################################################################################
#
#   PROCESS FILE 2 ไฟล์ GL ที่เป็น Excel อยู่แล้ว
#
###################################################################################

    if filename2.cget('text') != "[ไม่ได้เลือก]":
        txtbox1.insert(tk.END, '[ประมวลผลไฟล์ GL]\n')
        txtbox1.yview_pickplace("end")
        glfilename = filename2.cget('text')
        #cols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
        # colsname = ['อ้างอิง', 'การกำหนด', 'เลขที่เอกสาร', 'วันที่เอกสาร', 'วันที่ผ่านรายการ', 'คีย์การผ่านรายการ',
        #            'ประเภทเอกสาร', 'ชื่อผู้ใช้', 'เขตธุรกิจ', 'จำนวนเงิน', 'เอกสารกาหักล้าง', 'วันที่หักล้าง', 'ข้อความ']
        colsname = ['การกำหนด', 'เลขที่เอกสาร', 'เขตธุรกิจ', 'วันที่เอกสาร', 'วันที่ผ่านรายการ', 'คีย์การผ่านรายการ',
                    'จำนวนเงิน', 'เอกสารการหักล้าง', 'วันที่หักล้าง', 'ข้อความ', 'การอ้างอิง']

        txtbox1.insert(tk.END, 'อ่านไฟล์ ' + glfilename + '\n')
        txtbox1.yview_pickplace("end")

        df2 = pd.read_excel(glfilename, names=colsname,
                            dtype={'เลขที่เอกสาร': 'str', 'จำนวนเงิน': 'str'})
        # กรองเอาเฉพาะ 'บันทึกชดเชยเงินสดย่อยจากรายได้'
        
        # มีบางสาขาบันทึกข้อความเป็น บันทึกชดเชยเงินสดย่อยชุดที่ xxx
        #df2 = df2[df2['ข้อความ'].str[0:30] == 'บันทึกชดเชยเงินสดย่อยจากรายได้']
        df2 = df2[df2['ข้อความ'].str[0:21] == 'บันทึกชดเชยเงินสดย่อย']

        # เพิ่มฟิล ปีเดือน เพื่อใช้ในการกรองงวด
        df2['ym'] = df2['วันที่ผ่านรายการ'].map(lambda x: x.strftime('%Y%m'))
        ym = df2.groupby('ym')['ym'].min()
        # วิเคราะห์ข้อมูล
        txtbox1.insert(tk.END, '[วิเคราะห์ข้อมูล]\n')
        txtbox1.yview_pickplace("end")
        for i in ym:
            df_temp = df2[df2['ym'] == i]
            tym = misc.thaiym(df_temp['วันที่ผ่านรายการ'].min())
            txtbox1.insert(tk.END, 'งวดเดือน [' + tym + ']\n')
            txtbox1.yview_pickplace("end")
            # จำนวนรายการ
            record_count = df_temp['วันที่ผ่านรายการ'].count()
            txtbox1.insert(tk.END, 'จำนวนรายการที่เป็น "บันทึกชดเชยเงินสดย่อยจากรายได้" : ' +
                           str(record_count) + ' รายการ\n')
            txtbox1.yview_pickplace("end")

            # ช่วงของวันที่ผ่านรายการ
            dmin = df_temp['วันที่ผ่านรายการ'].min()
            dmax = df_temp['วันที่ผ่านรายการ'].max()
            txtbox1.insert(tk.END, 'วันที่ผ่านรายการ ระหว่างวันที่ ' + misc.thaidate(dmin) +
                           ' ถึงวันที่ '+misc.thaidate(dmax) + '\n')
            txtbox1.yview_pickplace("end")
        # รวมฝั่งบวก และ ลบ
            positive_sum = Decimal('0')
            negative_sum = Decimal('0')
            positive_max = None
            positive_min = None
            for i in df_temp['จำนวนเงิน']:
                if Decimal(i) > 0:
                    positive_sum += Decimal(i)
                else:
                    negative_sum += Decimal(i)
                    if positive_max != None:
                        if abs(Decimal(i)) > positive_max:
                            positive_max = abs(Decimal(i))
                    else:
                        positive_max = abs(Decimal(i))
                    if positive_min != None:
                        if abs(Decimal(i)) < positive_min:
                            positive_min = abs(Decimal(i))
                    else:
                        positive_min = abs(Decimal(i))
            txtbox1.insert(
                tk.END, 'จำนวนเงินฝั่งบวก = {:,.2f} บาท\n'.format(positive_sum))
            txtbox1.yview_pickplace("end")
            txtbox1.insert(
                tk.END, 'จำนวนเงินฝั่งลบ = {:,.2f} บาท\n'.format(negative_sum))
            txtbox1.yview_pickplace("end")
            txtbox1.insert(tk.END, 'ผลต่าง {:,.2f} บาท\n'.format(
                positive_sum+negative_sum))
            txtbox1.yview_pickplace("end")
            gl_amount = abs(negative_sum)
            txtbox1.insert(
                tk.END, 'จำนวนเงินต่ำสุดคือ {:,.2f} บาท\n'.format(positive_min))
            txtbox1.yview_pickplace("end")
            txtbox1.insert(
                tk.END, 'จำนวนเงินสูงสุดคือ {:,.2f} บาท\n'.format(positive_max))
            txtbox1.yview_pickplace("end")
            df_temp['จำนวนเงิน'] = df_temp['จำนวนเงิน'].astype('float')
            with pd.ExcelWriter(result_filename, mode='a', engine='openpyxl') as writer:
                df_temp.to_excel(writer, sheet_name='GL '+tym, index=False)
                txtbox1.insert(tk.END, 'ส่งออก GL ไปยัง "' +
                               result_filename + '"\nSheet Name: ' + 'GL '+tym + '\n\n')
                txtbox1.yview_pickplace("end")

#######################################################################################
#
#          สรุป
#
#######################################################################################


    submit_btn['state'] = tk.NORMAL
    txtbox1.insert(tk.END, '[จบกระบวนการ]\n')
    txtbox1.yview_pickplace("end")


def btn1_command():
    f = fd.askopenfilename(filetypes=[("ZFARP08", ".xls")])
    if f != '':
        filename1.config(text=f)
    else:
        filename1.config(text='[ไม่ได้เลือก]')


def btn2_command():
    f = fd.askopenfilename(filetypes=[("GL Excel File", ".xlsx")])
    if f != '':
        filename2.config(text=f)
    else:
        filename2.config(text='[ไม่ได้เลือก]')


def process_command():
    submit_btn['state'] = tk.DISABLED
    thread1 = threading.Thread(target=process_file)

    thread1.start()

    # process_file1()
    # proc_window=tk.Toplevel(window)
    # proc_window.title="ประมวลผล"

## จัดหน้าตา UI
window = tk.Tk()
window.title("เครื่องมือตรวจสอบ การจ่ายเงินสดย่อย โดยสำนักตรวจสอบกระบวนการหลัก")
print(window.geometry())

s = screen_config.scaling(window.winfo_id())

window.tk.call('tk', 'scaling', s)


print(window.geometry())
upper_panel1 = tk.Frame(window, padx=2, pady=2)
upper_panel1.grid(row=0, column=0, sticky='we')
upper_panel2 = tk.Frame(window, padx=10, pady=2)
upper_panel2.grid(row=1, column=0, sticky='we')
upper_panel3 = tk.Frame(window, padx=2, pady=2)
upper_panel3.grid(row=2, column=0, sticky='we')
#lower_panel = tk.Frame(window, padx=5, pady=5,bg='cyan')
# lower_panel.grid(row=3,column=0,sticky='nsew')

tk.Label(upper_panel1, text="ไฟล์ รายงานการจ่ายเงินสดย่อย(ZFAPRP08) :").pack(side="left")
filename1 = tk.Label(upper_panel1, text="[ไม่ได้เลือก]")
filename1.pack(side='left')
tk.Button(upper_panel1, text="เลือกไฟล์",
          command=btn1_command).pack(side='left')
tk.Label(upper_panel2, text="ไฟล์ GL-เงินสด(FBL3N) :").pack(side='left')
filename2 = tk.Label(upper_panel2, text="[ไม่ได้เลือก]")
filename2.pack(side='left')
tk.Button(upper_panel2, text="เลือกไฟล์",
          command=btn2_command).pack(side="left")
submit_btn = tk.Button(upper_panel3, text="ดำเนินการ", command=process_command)
submit_btn.pack(side='left')
tk.Label(upper_panel3,text='เวอร์ชัน 1.05.00 [วันที่ 13 กุมภาพันธ์ 2567]',fg='#0000ff').pack()
# เวอร์ชัน 1.00.00 เป็นตัวแรก
#
# 6 ธันวาคม 2566
# เวอร์ชัน 1.02.00 ปรับแก้บัคเปิดไฟล์ ZFAPRP08
# 
# 26 มกราคม 2567
# เวอร์ชัน 1.03.00 แก้บัคจำนวน tab จากตัวอักษรไปหาจำนวนเงินมี 3 tab (โดยแทปรวมมี 21 แทป)
# 6 กุมภาพันธ์ 2567
# เวอร์ชัน 1.03.01 ปรับปรุงให้บันทึก result ไว้ directory เดียวกับไฟล์ข้อมูลเนื่องจาบัค onedrive จะบันทึกลง desktop ไม่ได้
# 7 กุมภาพันธ์ 2567
# เวอร์ชัน 1.04.00 แก้บัคเปิดไฟล์ ZFARP08 เปลี่ยนอัลกอริธึมใหม่
# 13 กุมภาพันธ์ 2567
# เวอร์ชัน 1.05.00 แก้บัค User ใส่ # ในรายละเอียด ทำให้ pandas ไม่ประมวลผลเนื่องจากได้กำหนด # เป็นคอมเม้น
# แก้ไขโดยเปลี่ยน comment เป็น ` แทนเหตุผลที่ใช้ตัวนี้คือ ปกติ windows ไม่สามารถพิมพ์อักษรตัวนี้ได้
txtbox1 = tk.Text(window, font=('Tahoma', 12))
#txtbox1 = tk.scrolledtext.ScrolledText(window)
#v=tk.Scrollbar(lower_panel, orient='vertical')
# v.pack(side='right',fill='y')
# v.config(command=txtbox1.yview)
txtbox1.grid(row=3, column=0, sticky='sew')
print(txtbox1.winfo_height)
# print(fd.askopenfile())

window.mainloop()
